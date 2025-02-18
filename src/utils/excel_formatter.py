"""
Módulo para el formateo de archivos Excel.
"""

import pandas as pd
from openpyxl.styles import Font
from openpyxl.styles.colors import BLUE

class ExcelFormatter:
    """Clase para manejar el formateo de archivos Excel."""

    def __init__(self, df, output_path, links=None):
        """
        Inicializa el formateador.
        
        Args:
            df (pandas.DataFrame): DataFrame a formatear
            output_path (str): Ruta del archivo de salida
            links (list): Lista de links para los hipervínculos
        """
        self.df = df
        self.output_path = output_path
        self.links = links or []
        self.writer = None
        
        # Convertir columnas de fecha a datetime
        self._convert_date_columns()

    
    def _convert_date_columns(self):
        """Convierte las columnas de fecha a formato datetime de pandas."""
        date_columns = [
            'Fecha Inicio',
            'Fecha Creación',
            'Fecha Actualización'
        ]
        
        for col in date_columns:
            if col in self.df.columns:
                # Convertir a datetime usando pandas
                self.df[col] = pd.to_datetime(
                    self.df[col], 
                    format='%d/%m/%Y', 
                    errors='coerce'
                )

    def format_excel(self):
        """
        Aplica el formato al archivo Excel.
        
        Returns:
            bool: True si el proceso fue exitoso
        """
        try:
            # Usar openpyxl como motor y especificar el formato de fecha
            self.writer = pd.ExcelWriter(
                self.output_path,
                engine='openpyxl',
                datetime_format='dd/mm/yyyy'
            )
            
            # Escribir el DataFrame al Excel
            self.df.to_excel(self.writer, index=False, sheet_name='Tareas')
            worksheet = self.writer.sheets['Tareas']
            
            self._format_worksheet(worksheet)
            self._add_hyperlinks(worksheet)
            self._format_columns(worksheet)
            self._adjust_columns(worksheet)
            
            self.writer.close()
            return True
        except Exception as e:
            print(f"Error formateando Excel: {str(e)}")
            return False

    def _format_columns(self, worksheet):
        """
        Aplica formato a las columnas según su tipo.
        
        Args:
            worksheet: Hoja de trabajo de Excel
        """
        # Formato para horas
        if 'Horas Utilizadas' in self.df.columns:
            col_idx = list(self.df.columns).index('Horas Utilizadas') + 1
            for row in range(2, len(self.df) + 2):
                cell = worksheet.cell(row=row, column=col_idx)
                cell.number_format = '#,##0.0'

        # Formato para fechas
        date_columns = {
            'Fecha Inicio': 'dd/mm/yyyy',
            'Fecha Creación': 'dd/mm/yyyy',
            'Fecha Actualización': 'dd/mm/yyyy'
        }

        for col_name, date_format in date_columns.items():
            if col_name in self.df.columns:
                col_idx = list(self.df.columns).index(col_name) + 1
                for row in range(2, len(self.df) + 2):
                    cell = worksheet.cell(row=row, column=col_idx)
                    if cell.value:  # Solo si hay valor
                        cell.number_format = date_format

        # Formato para horas (HH:MM:SS)
        time_columns = [
            'Hora Creación',
            'Hora Actualización'
        ]

        for col_name in time_columns:
            if col_name in self.df.columns:
                col_idx = list(self.df.columns).index(col_name) + 1
                for row in range(2, len(self.df) + 2):
                    cell = worksheet.cell(row=row, column=col_idx)
                    if cell.value:  # Solo si hay valor
                        cell.number_format = 'hh:mm:ss'
                        
    def _format_worksheet(self, worksheet):
        """
        Aplica formato general a la hoja de trabajo.
        
        Args:
            worksheet: Hoja de trabajo de Excel
        """
        # Aquí puedes añadir formato general a la hoja
        pass

    def _adjust_columns(self, worksheet):
        """
        Ajusta el ancho de las columnas.
        
        Args:
            worksheet: Hoja de trabajo de Excel
        """
        for idx, col in enumerate(self.df.columns):
            max_length = max(
                self.df[col].astype(str).apply(len).max(),
                len(col)
            )
            worksheet.column_dimensions[chr(65 + idx)].width = max_length + 2

    def _format_hours(self, worksheet):
        """
        Aplica formato a las columnas de horas y fechas.
        """
        # Formato para horas
        if 'Horas Utilizadas' in self.df.columns:
            col_idx = list(self.df.columns).index('Horas Utilizadas')
            for row in range(2, len(self.df) + 2):
                cell = worksheet.cell(row=row, column=col_idx + 1)
                cell.number_format = '#,##0.0'
        
        # Formato para fecha de inicio
        if 'Fecha Inicio' in self.df.columns:
            col_idx = list(self.df.columns).index('Fecha Inicio')
            for row in range(2, len(self.df) + 2):
                cell = worksheet.cell(row=row, column=col_idx + 1)
                if cell.value:  # Solo si hay valor
                    cell.number_format = 'dd/mm/yyyy'

    def _add_hyperlinks(self, worksheet):
        """
        Añade hipervínculos a los códigos de tarea.
        
        Args:
            worksheet: Hoja de trabajo de Excel
        """
        if self.links:
            for idx, (codigo, link) in enumerate(zip(self.df['Código'], self.links), start=2):
                if codigo and link:  # Verifica que ambos valores existan
                    cell = worksheet.cell(row=idx, column=1)
                    cell.hyperlink = link
                    cell.font = Font(color=BLUE, underline="single")
                    cell.value = codigo